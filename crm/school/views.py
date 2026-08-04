from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Avg, Q
from django.utils import timezone
from django.urls import reverse
from .decorators import admin_required, admin_or_teacher, all_roles, role_required
from .models import (
    Student, Teacher, Classroom, Grade, Subject,
    Attendance, TeacherAttendance, Score, AcademicYear, ExamType, Exam,
    Timetable, TimeSlot, Notification, NotificationRead,
    ReportCard, SchoolEvent, UserProfile, SchoolSettings
)
from .forms import (
    StudentForm, TeacherForm, ClassroomForm, AttendanceForm,
    BulkAttendanceForm, TeacherAttendanceForm, BulkTeacherAttendanceForm,
    ScoreForm, SubjectForm, GradeForm,
    AcademicYearForm, ExamTypeForm, ExamForm, TimetableForm,
    TimeSlotForm, NotificationForm, ReportCardForm,
    SchoolEventForm, LoginForm, UserCreateForm, ProfileUpdateForm,
    SchoolSettingsForm, StudentRegisterForm, ParentRegisterForm
)

# ══════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════
def login_view(request):
    if request.user.is_authenticated:
        return redirect('school:dashboard')
    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        login(request, user)
        
        # Track login with device information
        from .models import LoginHistory
        from .utils import get_client_ip, parse_user_agent, create_login_notification
        
        device_info = parse_user_agent(request)
        ip_address = get_client_ip(request)
        
        # Create login history record
        login_history = LoginHistory.objects.create(
            user=user,
            ip_address=ip_address,
            device_type=device_info['device_type'],
            browser=device_info['browser'],
            operating_system=device_info['operating_system'],
            device_name=device_info['device_name'],
            user_agent=device_info['user_agent'],
        )
        
        # Create notification for admins
        create_login_notification(user, login_history)
        
        name = user.get_full_name() or user.username
        messages.success(request, f'សូមស្វាគមន៍, {name}!')
        return redirect(request.GET.get('next', 'school:dashboard'))
    return render(request, 'school/auth/login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, 'អ្នកបានចាកចេញពីប្រព័ន្ធ។')
    return redirect('school:login')


# ══════════════════════════════════════════════
#  REGISTER — Student
# ══════════════════════════════════════════════
def register_student(request):
    """Public view: a student creates their own login account."""
    if request.user.is_authenticated:
        return redirect('school:dashboard')
    form = StudentRegisterForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(request, f'ស្វាគមន៍, {user.get_full_name() or user.username}! គណនីសិស្សត្រូវបានបង្កើតដោយជោគជ័យ។')
        return redirect('school:dashboard')
    return render(request, 'school/auth/register_student.html', {'form': form})


# ══════════════════════════════════════════════
#  REGISTER — Parent (Mom / Dad)
# ══════════════════════════════════════════════
def register_parent(request):
    """Public view: a mom or dad creates their own login account."""
    if request.user.is_authenticated:
        return redirect('school:dashboard')
    form = ParentRegisterForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        try:
            user = form.save()
            login(request, user)
            messages.success(request, f'ស្វាគមន៍, {user.get_full_name() or user.username}! គណនីមាតាបិតាត្រូវបានបង្កើតដោយជោគជ័យ។')
            return redirect('school:dashboard')
        except Exception as e:
            messages.error(request, f'កំហុសក្នុងការបង្កើតគណនី: {str(e)}')
    return render(request, 'school/auth/register_parent.html', {'form': form})


# ══════════════════════════════════════════════
#  PROFILE UPDATE (all roles)
# ══════════════════════════════════════════════
@login_required
def profile_update(request):
    profile = request.user.profile
    form = ProfileUpdateForm(
        request.POST or None,
        request.FILES or None,
        instance=profile,
        user=request.user
    )
    if request.method == 'POST' and form.is_valid():
        try:
            form.save_user(request.user)
            form.save()
            messages.success(request, 'ប្រវត្តិរូបបានធ្វើបច្ចុប្បន្នភាព។')
            return redirect('school:profile_update')
        except Exception as e:
            import traceback, logging
            logging.getLogger(__name__).error("profile_update failed: %s\n%s", e, traceback.format_exc())
            messages.error(request, f'បញ្ហា: {e}')
    return render(request, 'school/auth/profile.html', {
        'form': form, 'profile': profile
    })


# ══════════════════════════════════════════════
#  SCHOOL SETTINGS (Admin only)
# ══════════════════════════════════════════════
@admin_required
def school_settings_view(request):
    settings_obj = SchoolSettings.get()
    form = SchoolSettingsForm(
        request.POST or None,
        request.FILES or None,
        instance=settings_obj
    )
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'ការកំណត់សាលាបានរក្សាទុក។')
                return redirect('school:school_settings')
            except Exception as e:
                import traceback, logging
                logging.getLogger(__name__).error(
                    "school_settings_view save failed: %s\n%s",
                    e, traceback.format_exc()
                )
                messages.error(request, f'មានបញ្ហាក្នុងការរក្សាទុក: {e}')
        # If form invalid or save failed, fall through to re-render with errors
    return render(request, 'school/school_settings.html', {
        'form': form, 'settings': settings_obj
    })

# ══════════════════════════════════════════════
#  DASHBOARD — role-specific
# ══════════════════════════════════════════════
@login_required
def dashboard(request):
    today = timezone.now().date()
    try:
        role = request.user.profile.role
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Profile access error in dashboard: {e}", exc_info=True)
        role = 'student'

    notifications  = Notification.objects.filter(is_active=True).order_by('-created_at')[:5]
    upcoming_events = SchoolEvent.objects.filter(start_date__gte=today).order_by('start_date')[:5]

    # ── ADMIN dashboard ──────────────────────
    if role == 'admin':
        return render(request, 'school/dashboard.html', {
            'role': role,
            'total_students':   Student.objects.filter(is_active=True).count(),
            'total_teachers':   Teacher.objects.filter(is_active=True).count(),
            'total_classrooms': Classroom.objects.count(),
            'total_subjects':   Subject.objects.count(),
            'today_present': Attendance.objects.filter(date=today, status='P').count(),
            'today_absent':  Attendance.objects.filter(date=today, status='A').count(),
            'today_late':    Attendance.objects.filter(date=today, status='L').count(),
            'recent_students': Student.objects.filter(is_active=True).order_by('-enrolled_date')[:5],
            'recent_scores':   Score.objects.select_related('student','subject','exam_type').order_by('-date_recorded')[:5],
            'notifications': notifications,
            'upcoming_events': upcoming_events,
            'today': today,
        })

    # ── TEACHER dashboard ────────────────────
    if role == 'teacher':
        try:
            teacher = request.user.profile.teacher
        except Exception:
            teacher = None
        my_classes   = Classroom.objects.filter(homeroom_teacher=teacher) if teacher else Classroom.objects.none()
        my_students  = Student.objects.filter(classroom__in=my_classes, is_active=True) if teacher else Student.objects.none()
        today_att    = Attendance.objects.filter(date=today, student__in=my_students)
        return render(request, 'school/dashboard_teacher.html', {
            'role': role, 'teacher': teacher,
            'my_classes': my_classes,
            'my_students_count': my_students.count(),
            'today_present': today_att.filter(status='P').count(),
            'today_absent':  today_att.filter(status='A').count(),
            'today_late':    today_att.filter(status='L').count(),
            'recent_scores': Score.objects.filter(student__in=my_students).select_related('student','subject','exam_type').order_by('-date_recorded')[:5],
            'notifications': notifications,
            'upcoming_events': upcoming_events,
            'today': today,
        })

    # ── PARENT dashboard ─────────────────────
    if role == 'parent':
        try:
            student = request.user.profile.student
        except Exception:
            student = None
        att = student.attendances.order_by('-date')[:10] if student else []
        scores = student.scores.select_related('subject','exam_type').order_by('-date_recorded')[:10] if student else []
        return render(request, 'school/dashboard_parent.html', {
            'role': role, 'student': student,
            'attendances': att, 'scores': scores,
            'notifications': notifications,
            'upcoming_events': upcoming_events,
            'today': today,
        })

    # ── STUDENT dashboard ────────────────────
    try:
        student = request.user.profile.student
    except Exception:
        student = None
    att    = student.attendances.order_by('-date')[:5] if student else []
    scores = student.scores.select_related('subject','exam_type').order_by('-date_recorded')[:5] if student else []
    timetables = []
    if student and student.classroom:
        timetables = Timetable.objects.filter(
            classroom=student.classroom
        ).select_related('subject','teacher','time_slot').order_by('time_slot__day','time_slot__start_time')
    return render(request, 'school/dashboard_student.html', {
        'role': role, 'student': student,
        'attendances': att, 'scores': scores,
        'timetables': timetables,
        'notifications': notifications,
        'upcoming_events': upcoming_events,
        'today': today,
    })

# ══════════════════════════════════════════════
#  USER MANAGEMENT (Admin only)
# ══════════════════════════════════════════════
@admin_required
def user_list(request):
    # select_related avoids N+1 queries; profiles are guaranteed by the
    # ensure_user_profile context processor and post_save signal.
    users = User.objects.select_related('profile').order_by('username')
    # Safety net: create any missing profiles in a single pass
    missing = [u for u in users if not hasattr(u, 'profile') or u.profile is None]
    if missing:
        for u in missing:
            role = 'admin' if u.is_superuser else 'student'
            UserProfile.objects.get_or_create(user=u, defaults={'role': role})
        users = User.objects.select_related('profile').order_by('username')
    return render(request, 'school/users/user_list.html', {'users': users})

@admin_required
def user_add(request):
    form = UserCreateForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'អ្នកប្រើបានបង្កើតរួច។')
        return redirect('school:user_list')
    return render(request, 'school/users/user_add.html', {'form': form})

@admin_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'អ្នកប្រើបានលុបរួច។')
        return redirect('school:user_list')
    return render(request, 'school/confirm_delete.html', {
        'object': user, 'title': 'លុបអ្នកប្រើ', 'back_url': reverse('school:user_list')
    })

# ══════════════════════════════════════════════
#  STUDENTS (Admin: full CRUD | Teacher: view | Parent/Student: own only)
# ══════════════════════════════════════════════
@admin_or_teacher
def student_list(request):
    try:
        role = request.user.profile.role
    except Exception:
        role = 'admin' if (request.user.is_superuser or request.user.is_staff) else 'teacher'
    q = request.GET.get('q', '')
    classroom_id = request.GET.get('classroom', '')
    status_filter = request.GET.get('status', '')  # New status filter
    
    students = Student.objects.filter(is_active=True).select_related('classroom__grade')
    
    # Teacher sees only their class students
    if role == 'teacher':
        try:
            teacher = request.user.profile.teacher
            my_classes = Classroom.objects.filter(homeroom_teacher=teacher)
            students = students.filter(classroom__in=my_classes)
        except Exception:
            students = Student.objects.none()
    
    if q:
        students = students.filter(Q(first_name__icontains=q)|Q(last_name__icontains=q)|Q(student_id__icontains=q))
    if classroom_id:
        students = students.filter(classroom_id=classroom_id)
    if status_filter:
        students = students.filter(status=status_filter)
    
    classrooms = Classroom.objects.select_related('grade','academic_year')
    
    # Get status choices for filter dropdown
    status_choices = Student.STATUS_CHOICES
    
    return render(request, 'school/student_list.html', {
        'students': students, 'q': q,
        'classrooms': classrooms, 'selected_classroom': classroom_id, 'role': role,
        'status_choices': status_choices, 'selected_status': status_filter,
    })



@admin_or_teacher
def student_list_export_excel(request):
    """
    Export student list to Excel file with comprehensive information
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get filter parameters (same as student_list)
    try:
        role = request.user.profile.role
    except Exception:
        role = 'admin' if (request.user.is_superuser or request.user.is_staff) else 'teacher'
    
    q = request.GET.get('q', '')
    classroom_id = request.GET.get('classroom', '')
    students = Student.objects.filter(is_active=True).select_related('classroom__grade', 'classroom__academic_year')
    
    # Teacher sees only their class students
    if role == 'teacher':
        try:
            teacher = request.user.profile.teacher
            my_classes = Classroom.objects.filter(homeroom_teacher=teacher)
            students = students.filter(classroom__in=my_classes)
        except Exception:
            students = Student.objects.none()
    
    if q:
        students = students.filter(Q(first_name__icontains=q)|Q(last_name__icontains=q)|Q(student_id__icontains=q))
    if classroom_id:
        students = students.filter(classroom_id=classroom_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers (English and Khmer)
    headers = [
        'Student ID\nលេខសម្គាល់',
        'Last Name (KH)\nនាមត្រកូល',
        'First Name (KH)\nនាមខ្លួន',
        'Last Name (EN)',
        'First Name (EN)',
        'Gender\nភេទ',
        'Date of Birth\nថ្ងៃខែឆ្នាំកំណើត',
        'Place of Birth\nទីកន្លែងកំណើត',
        'Nationality\nសញ្ជាតិ',
        'Religion\nសាសនា',
        'Birth Cert No.\nលេខសំបុត្រកំណើត',
        'Classroom\nថ្នាក់រៀន',
        'Phone\nទូរស័ព្ទ',
        'Address\nអាសយដ្ឋាន',
        'Father Name\nឈ្មោះឪពុក',
        'Father Phone\nទូរស័ព្ទឪពុក',
        'Father Occupation\nមុខរបរឪពុក',
        'Mother Name\nឈ្មោះម្តាយ',
        'Mother Phone\nទូរស័ព្ទម្តាយ',
        'Mother Occupation\nមុខរបរម្តាយ',
        'Guardian Name\nឈ្មោះអាណាព្យាបាល',
        'Guardian Phone\nទូរស័ព្ទអាណាព្យាបាល',
        'Guardian Email\nអ៊ីម៉ែលអាណាព្យាបាល',
        'Emergency Contact\nទំនាក់ទំនងបន្ទាន់',
        'Emergency Phone\nទូរស័ព្ទបន្ទាន់',
        'Blood Group\nក្រុមឈាម',
        'Allergies\nអាឡែកស៊ី',
        'Enrolled Date\nថ្ងៃចុះឈ្មោះ',
        'Previous School\nសាលារៀនមុន',
        'Status\nស្ថានភាព',
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Set row height for header
    ws.row_dimensions[1].height = 40
    
    # Write data
    for row_num, student in enumerate(students, 2):
        data = [
            student.student_id or '',
            student.last_name or '',
            student.first_name or '',
            student.last_name_en or '',
            student.first_name_en or '',
            'Male' if student.gender == 'M' else 'Female' if student.gender == 'F' else '',
            student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '',
            student.place_of_birth or '',
            student.nationality or '',
            student.religion or '',
            student.birth_certificate_number or '',
            str(student.classroom) if student.classroom else '',
            student.phone or '',
            student.address or '',
            student.father_name or '',
            student.father_phone or '',
            student.father_occupation or '',
            student.mother_name or '',
            student.mother_phone or '',
            student.mother_occupation or '',
            student.parent_name or '',
            student.parent_phone or '',
            student.parent_email or '',
            student.emergency_contact_name or '',
            student.emergency_contact_phone or '',
            student.blood_group or '',
            student.allergies or '',
            student.enrolled_date.strftime('%d/%m/%Y') if student.enrolled_date else '',
            student.previous_school or '',
            'Active' if student.is_active else 'Inactive',
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'students_list_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@admin_or_teacher
@admin_or_teacher
def student_detail(request, pk):
    from django.db.models import Avg
    from .models import StudentHistory
    
    student = get_object_or_404(Student, pk=pk)
    attendances   = student.attendances.order_by('-date')[:10]
    scores        = student.scores.select_related('subject','exam_type','academic_year').order_by('-date_recorded')
    
    # Get student history records (previous grades)
    history_records = student.history_records.select_related('academic_year', 'classroom').order_by('-academic_year__year')
    
    # Calculate statistics
    present_count = student.attendances.filter(status='P').count()
    absent_count  = student.attendances.filter(status='A').count()
    total_attendance = student.attendances.count()
    attendance_rate = round((present_count / total_attendance * 100), 1) if total_attendance > 0 else 0
    
    # Calculate average score
    average_score = scores.aggregate(avg=Avg('score'))['avg']
    if average_score:
        average_score = round(average_score, 1)
    
    return render(request, 'school/student_detail.html', {
        'student': student, 
        'attendances': attendances, 
        'scores': scores,
        'history_records': history_records,
        'present_count': present_count, 
        'absent_count': absent_count,
        'attendance_rate': attendance_rate,
        'average_score': average_score,
        'role': getattr(request.user.profile, 'role', 'student') if hasattr(request.user, 'profile') else 'student',
    })

@admin_required
def student_add(request):
    form = StudentForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, 'សិស្សបានបន្ថែមរួច។')
            return redirect('school:student_list')
        except Exception as e:
            import traceback, logging
            logging.getLogger(__name__).error("student_add failed: %s\n%s", e, traceback.format_exc())
            messages.error(request, f'បញ្ហា: {e}')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'បន្ថែមសិស្ស', 'back_url': reverse('school:student_list')
    })

@admin_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form = StudentForm(request.POST or None, request.FILES or None, instance=student)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, 'សិស្សបានកែប្រែ។')
            return redirect('school:student_detail', pk=pk)
        except Exception as e:
            import traceback, logging
            logging.getLogger(__name__).error("student_edit failed: %s\n%s", e, traceback.format_exc())
            messages.error(request, f'បញ្ហា: {e}')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'កែប្រែសិស្ស',
        'subtitle': f'ID: {student.student_id}',
        'back_url': reverse('school:student_list'),
    })

@admin_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.is_active = False
        student.save()
        messages.success(request, 'សិស្សបានដកចេញ។')
        return redirect('school:student_list')
    return render(request, 'school/confirm_delete.html', {
        'object': student, 'title': 'ដកសិស្ស', 'back_url': reverse('school:student_list')
    })


@admin_or_teacher
def student_history(request, pk):
    """
    Display complete academic history for a student
    បង្ហាញប្រវត្តិសិក្សាពេញលេញរបស់សិស្ស
    """
    from django.db.models import Count, Avg
    from .models import StudentHistory
    
    student = get_object_or_404(Student, pk=pk)
    
    # Get all history records ordered by year
    history_records = student.history_records.select_related(
        'academic_year', 'classroom', 'classroom__grade'
    ).order_by('-academic_year__year')
    
    # Add additional stats for each history record
    history_data = []
    for record in history_records:
        history_data.append({
            'record': record,
            'attendance_percentage': record.attendance_percentage(),
            'pass_percentage': record.pass_percentage(),
            'status_icon': '✅' if record.status == 'PROMOTED' else '🔄' if record.status == 'ACTIVE' else '❌',
            'status_color': 'success' if record.status == 'PROMOTED' else 'primary' if record.status == 'ACTIVE' else 'danger',
        })
    
    # Overall statistics
    total_years = history_records.count()
    total_promoted = history_records.filter(status='PROMOTED').count()
    overall_avg = history_records.aggregate(avg=Avg('average_score'))['avg']
    
    # Current year info
    current_info = {
        'classroom': student.classroom,
        'status': student.status,
        'year': student.classroom.academic_year if student.classroom else None,
    }
    
    return render(request, 'school/student_history.html', {
        'student': student,
        'history_data': history_data,
        'current_info': current_info,
        'total_years': total_years,
        'total_promoted': total_promoted,
        'overall_avg': round(overall_avg, 1) if overall_avg else 0,
        'role': getattr(request.user.profile, 'role', 'student') if hasattr(request.user, 'profile') else 'student',
    })


@admin_or_teacher
def student_promote(request):
    """
    Bulk promote students to next grade based on Cambodia Education System
    ដាក់សិស្សឡើងថ្នាក់ជាក្រុមតាមប្រព័ន្ធអប់រំកម្ពុជា
    
    Cambodia Education System Standards:
    - Primary (បឋមសិក្សា): Grade 1-6
    - Lower Secondary (បឋមភូមិ): Grade 7-9
    - Upper Secondary (មធ្យមភូមិ): Grade 10-12
    
    Promotion Criteria (លក្ខខណ្ឌឡើងថ្នាក់):
    1. Average score ≥ 50% across all subjects (ពិន្ទុមធ្យម ≥ 50%)
    2. Must have at least 1 subject with scores (ត្រូវមានពិន្ទុយ៉ាងហោចណាស់ 1 មុខវិជ្ជា)
    3. Attendance rate ≥ 80% is recommended (វត្តមាន ≥ 80% ត្រូវបានផ្តល់អនុសាសន៍)
    4. Can only promote to immediate next grade (ឡើងបានតែថ្នាក់បន្ទាប់ប៉ុណ្ណោះ)
    
    Special Level Transitions (ការផ្ទេរកម្រិតពិសេស):
    - Grade 6 → Grade 7: Primary to Lower Secondary (បឋមសិក្សា → បឋមភូមិ)
    - Grade 9 → Grade 10: Lower Secondary to Upper Secondary (បឋមភូមិ → មធ្យមភូមិ)
    - Grade 12: Graduation (បញ្ចប់ការសិក្សា)
    
    Creates historical records to preserve student data across academic years.
    """
    from django.db.models import Count, Avg, Q
    from .models import StudentHistory
    
    # Get filter parameters
    current_classroom_id = request.GET.get('classroom', '')
    academic_year_id = request.GET.get('academic_year', '')
    passing_percentage = float(request.GET.get('passing_percentage', 50))
    
    classrooms = Classroom.objects.select_related('grade', 'academic_year')
    academic_years = AcademicYear.objects.all()
    
    students_data = []
    
    # Show students if classroom is selected (academic year is now optional)
    if current_classroom_id:
        current_classroom = get_object_or_404(Classroom, pk=current_classroom_id)
        
        # Get all students in the classroom
        students = Student.objects.filter(
            classroom=current_classroom,
            is_active=True
        ).prefetch_related('scores')
        
        for student in students:
            # Get scores - filter by academic year if provided
            if academic_year_id:
                academic_year = get_object_or_404(AcademicYear, pk=academic_year_id)
                scores = student.scores.filter(academic_year=academic_year)
            else:
                # Use current classroom's academic year or all scores
                if current_classroom.academic_year:
                    scores = student.scores.filter(academic_year=current_classroom.academic_year)
                else:
                    scores = student.scores.all()
            
            if scores.exists():
                # Calculate average percentage
                total_subjects = scores.count()
                avg_percentage = sum(score.percentage() for score in scores) / total_subjects if total_subjects > 0 else 0
                
                # Calculate attendance rate (វត្តមាន)
                if academic_year_id:
                    year = get_object_or_404(AcademicYear, pk=academic_year_id)
                    try:
                        if '-' in year.year:
                            start_year = year.year.split('-')[0]
                            end_year = year.year.split('-')[1]
                            year_attendance = student.attendances.filter(
                                date__gte=f'{start_year}-01-01',
                                date__lte=f'{end_year}-12-31'
                            )
                        else:
                            year_attendance = student.attendances.all()
                    except:
                        year_attendance = student.attendances.all()
                elif current_classroom.academic_year:
                    year = current_classroom.academic_year
                    try:
                        if '-' in year.year:
                            start_year = year.year.split('-')[0]
                            end_year = year.year.split('-')[1]
                            year_attendance = student.attendances.filter(
                                date__gte=f'{start_year}-01-01',
                                date__lte=f'{end_year}-12-31'
                            )
                        else:
                            year_attendance = student.attendances.all()
                    except:
                        year_attendance = student.attendances.all()
                else:
                    year_attendance = student.attendances.all()
                
                total_days = year_attendance.count()
                present_days = year_attendance.filter(status='P').count()
                attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0
                
                # Cambodia Education System Promotion Criteria:
                # លក្ខខណ្ឌឡើងថ្នាក់តាមប្រព័ន្ធអប់រំកម្ពុជា:
                # 1. ពិន្ទុមធ្យម >= passing_percentage (default 50%)
                # 2. ត្រូវមានយ៉ាងហោចណាស់ 1 មុខវិជ្ជា
                # 3. វត្តមាន >= 80% (អនុសាសន៍)
                
                can_promote = (
                    avg_percentage >= passing_percentage and 
                    total_subjects > 0 and
                    attendance_rate >= 80.0  # Attendance requirement
                )
                
                # Also calculate individual subject pass/fail for display
                passed_subjects = sum(1 for score in scores if score.is_passing(passing_percentage))
                failed_subjects = total_subjects - passed_subjects
                
                students_data.append({
                    'student': student,
                    'total_subjects': total_subjects,
                    'passed_subjects': passed_subjects,
                    'failed_subjects': failed_subjects,
                    'avg_percentage': round(avg_percentage, 1),
                    'attendance_rate': round(attendance_rate, 1),
                    'total_days': total_days,
                    'present_days': present_days,
                    'can_promote': can_promote,
                })
            else:
                # No scores - cannot promote
                students_data.append({
                    'student': student,
                    'total_subjects': 0,
                    'passed_subjects': 0,
                    'failed_subjects': 0,
                    'avg_percentage': 0,
                    'attendance_rate': 0,
                    'total_days': 0,
                    'present_days': 0,
                    'can_promote': False,
                })
    
    # Handle POST - actually promote students
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        next_classroom_id = request.POST.get('next_classroom')
        
        if student_ids and next_classroom_id:
            from django.utils import timezone
            next_classroom = get_object_or_404(Classroom, pk=next_classroom_id)
            promoted_count = 0
            failed_promotions = []
            
            for student_id in student_ids:
                try:
                    student = Student.objects.get(pk=student_id)
                    old_classroom = student.classroom
                    
                    if not old_classroom:
                        failed_promotions.append(f"{student}: មិនមានថ្នាក់បច្ចុប្បន្ន")
                        continue
                    
                    # Get grade information
                    old_grade = old_classroom.grade
                    new_grade = next_classroom.grade
                    old_grade_number = old_grade.grade_number if old_grade else 0
                    new_grade_number = new_grade.grade_number if new_grade else 0
                    
                    # VALIDATION 1: Must promote to next grade only (strict progression)
                    # មិនអនុញ្ញាតឱ្យរំលងថ្នាក់
                    if new_grade_number != old_grade_number + 1:
                        failed_promotions.append(
                            f"{student}: មិនអាចរំលងថ្នាក់បានទេ (ថ្នាក់ {old_grade_number} → ថ្នាក់ {new_grade_number})"
                        )
                        continue
                    
                    # VALIDATION 2: Check level transitions
                    # ពិនិត្យការផ្ទេរកម្រិត
                    if old_grade:
                        old_level = old_grade.level
                        new_level = new_grade.level if new_grade else ''
                        
                        # Verify correct level transitions
                        if old_level == 'primary' and old_grade_number == 6:
                            if new_level != 'lower_secondary' or new_grade_number != 7:
                                failed_promotions.append(
                                    f"{student}: ត្រូវផ្ទេរពីបឋមសិក្សាទៅបឋមភូមិ (Grade 6 → Grade 7)"
                                )
                                continue
                        
                        elif old_level == 'lower_secondary' and old_grade_number == 9:
                            if new_level != 'upper_secondary' or new_grade_number != 10:
                                failed_promotions.append(
                                    f"{student}: ត្រូវផ្ទេរពីបឋមភូមិទៅមធ្យមភូមិ (Grade 9 → Grade 10)"
                                )
                                continue
                        
                        # No promotion beyond Grade 12
                        elif old_grade_number == 12:
                            failed_promotions.append(
                                f"{student}: បញ្ចប់ការសិក្សាហើយ (Grade 12)"
                            )
                            continue
                    
                    # === CREATE HISTORY RECORD ===
                    # Save current academic year data before promotion
                    if old_classroom.academic_year:
                        # Get academic year data
                        year = old_classroom.academic_year
                        
                        # Calculate scores for this academic year
                        year_scores = student.scores.filter(academic_year=year)
                        total_subjects = year_scores.count()
                        if total_subjects > 0:
                            avg_score = sum(s.score for s in year_scores) / total_subjects
                            passed = sum(1 for s in year_scores if s.is_passing(passing_percentage))
                            failed = total_subjects - passed
                        else:
                            avg_score = 0
                            passed = 0
                            failed = 0
                        
                        # Calculate attendance for this academic year
                        # Parse year string (e.g., "2024-2025")
                        try:
                            if '-' in year.year:
                                start_year = year.year.split('-')[0]
                                end_year = year.year.split('-')[1]
                                year_attendance = student.attendances.filter(
                                    date__gte=f'{start_year}-01-01',
                                    date__lte=f'{end_year}-12-31'
                                )
                            else:
                                year_attendance = student.attendances.all()
                        except:
                            year_attendance = student.attendances.all()
                        
                        total_days = year_attendance.count()
                        present_days = year_attendance.filter(status='P').count()
                        absent_days = year_attendance.filter(status='A').count()
                        attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0
                        
                        # Determine next level transition
                        level_transition_note = ""
                        if old_grade_number == 6 and new_grade_number == 7:
                            level_transition_note = " | ✅ ផ្ទេរពីបឋមសិក្សាទៅបឋមភូមិ (Primary → Lower Secondary)"
                        elif old_grade_number == 9 and new_grade_number == 10:
                            level_transition_note = " | ✅ ផ្ទេរពីបឋមភូមិទៅមធ្យមភូមិ (Lower Secondary → Upper Secondary)"
                        elif old_grade_number == 12:
                            level_transition_note = " | 🎓 បញ្ចប់ការសិក្សា (Graduated)"
                        
                        # Create or update history record
                        history, created = StudentHistory.objects.update_or_create(
                            student=student,
                            academic_year=year,
                            defaults={
                                'classroom': old_classroom,
                                'grade_name': str(old_classroom.grade),
                                'grade_number': old_grade_number,
                                'grade_level': old_grade.level if old_grade else 'primary',
                                'status': 'PROMOTED',
                                'average_score': avg_score,
                                'total_subjects': total_subjects,
                                'passed_subjects': passed,
                                'failed_subjects': failed,
                                'total_days': total_days,
                                'present_days': present_days,
                                'absent_days': absent_days,
                                'end_date': timezone.now().date(),
                                'promoted_to': str(next_classroom),
                                'promotion_note': f"ឡើងថ្នាក់ទៅ {next_classroom.grade} នៅថ្ងៃទី {timezone.now().strftime('%d/%m/%Y')}{level_transition_note}",
                                'notes': f"ពិន្ទុមធ្យម: {avg_score:.1f} | វត្តមាន: {present_days}/{total_days} ថ្ងៃ ({attendance_rate:.1f}%)"
                            }
                        )
                    
                    # === UPDATE STUDENT RECORD ===
                    old_classroom_str = str(old_classroom)
                    student.previous_classroom = old_classroom_str
                    student.promotion_date = timezone.now().date()
                    
                    # Keep status as ACTIVE (they're active in new grade)
                    student.status = 'ACTIVE'
                    
                    # Move to new classroom
                    student.classroom = next_classroom
                    
                    # Add note with level transition info
                    level_note = ""
                    if old_grade_number == 6 and new_grade_number == 7:
                        level_note = " (✅ ចូលបឋមភូមិ)"
                    elif old_grade_number == 9 and new_grade_number == 10:
                        level_note = " (✅ ចូលមធ្យមភូមិ)"
                    
                    promotion_note = f"ឡើងថ្នាក់ពី {old_classroom_str} ទៅ {next_classroom} នៅថ្ងៃទី {timezone.now().strftime('%d/%m/%Y')}{level_note}"
                    if student.notes:
                        student.notes += f"\n{promotion_note}"
                    else:
                        student.notes = promotion_note
                    
                    student.save()
                    promoted_count += 1
                    
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error(f"Failed to promote student {student_id}: {e}")
                    failed_promotions.append(f"Student ID {student_id}: {str(e)}")
            
            # Show results
            if promoted_count > 0:
                messages.success(
                    request, 
                    f'✅ បានដាក់សិស្ស {promoted_count} នាក់ឡើងថ្នាក់ទៅ {next_classroom}។ ប្រវត្តិត្រូវបានរក្សាទុក។'
                )
            
            if failed_promotions:
                messages.warning(
                    request,
                    f'⚠️ មិនអាចដាក់ឡើងថ្នាក់បាន {len(failed_promotions)} នាក់:<br>' + 
                    '<br>'.join(failed_promotions)
                )
            
            return redirect('school:student_list')
    
    # Get available next grade classrooms with timetable info
    next_classrooms = []
    next_classrooms_with_timetable_info = []
    if current_classroom_id:
        current_classroom = Classroom.objects.get(pk=current_classroom_id)
        current_grade = current_classroom.grade
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"DEBUG: Current classroom: {current_classroom}, Grade: {current_grade}")
        
        if current_grade and current_grade.grade_number:
            current_grade_num = current_grade.grade_number
            logger.info(f"DEBUG: Looking for Grade {current_grade_num + 1} classrooms")
            
            # Get classrooms with next grade number
            all_classrooms = Classroom.objects.all().select_related('grade', 'academic_year')
            logger.info(f"DEBUG: Total classrooms in DB: {all_classrooms.count()}")
            
            for classroom in all_classrooms:
                if classroom.grade and classroom.grade.grade_number:
                    logger.info(f"DEBUG: Checking {classroom} - Grade {classroom.grade.grade_number}")
                    # Allow promotion to next grade only (strict progression)
                    if classroom.grade.grade_number == current_grade_num + 1:
                        next_classrooms.append(classroom)
                        logger.info(f"DEBUG: ✅ ADDED {classroom} to next classrooms")
                        
                        # Check if classroom has timetable
                        has_timetable = classroom.timetables.exists()
                        timetable_count = classroom.timetables.count()
                        next_classrooms_with_timetable_info.append({
                            'classroom': classroom,
                            'has_timetable': has_timetable,
                            'timetable_count': timetable_count
                        })
            
            logger.info(f"DEBUG: Final next_classrooms count: {len(next_classrooms_with_timetable_info)}")
        else:
            logger.warning(f"DEBUG: No grade or grade_number for {current_classroom}")
    
    return render(request, 'school/student_promote.html', {
        'classrooms': classrooms,
        'academic_years': academic_years,
        'students_data': students_data,
        'current_classroom_id': current_classroom_id,
        'current_classroom': Classroom.objects.get(pk=current_classroom_id) if current_classroom_id else None,
        'academic_year_id': academic_year_id,
        'passing_percentage': passing_percentage,
        'next_classrooms': next_classrooms,
        'next_classrooms_with_info': next_classrooms_with_timetable_info,
    })


# ══════════════════════════════════════════════
#  TEACHERS (Admin: full CRUD | Teacher: view self)
# ══════════════════════════════════════════════
@admin_or_teacher
def teacher_list(request):
    q = request.GET.get('q', '')
    teachers = Teacher.objects.filter(is_active=True)
    if q:
        teachers = teachers.filter(Q(first_name__icontains=q)|Q(last_name__icontains=q)|Q(subject_specialty__icontains=q))
    return render(request, 'school/teacher_list.html', {
        'teachers': teachers, 'q': q, 'role': getattr(request.user.profile, 'role', 'admin') if hasattr(request.user, 'profile') else 'admin'
    })

@admin_or_teacher
def teacher_detail(request, pk):
    from datetime import datetime, timedelta
    from django.db.models import Count, Q
    
    teacher = get_object_or_404(Teacher, pk=pk)
    subjects = teacher.subjects.all()
    classes = teacher.homeroom_classes.select_related('grade','academic_year').prefetch_related('students')
    
    # Attendance data
    attendances = teacher.attendances.order_by('-date')[:10]
    all_attendances = teacher.attendances.order_by('-date')
    
    # Timetable data
    timetables = teacher.timetables.select_related('time_slot','subject','classroom').order_by('time_slot__day','time_slot__start_time')
    
    # Calculate attendance statistics
    present_count = teacher.attendances.filter(status='P').count()
    total_attendance = teacher.attendances.count()
    attendance_rate = round((present_count / total_attendance * 100), 1) if total_attendance > 0 else 0
    
    # Monthly attendance (current month)
    today = datetime.now()
    month_start = today.replace(day=1)
    monthly_present = teacher.attendances.filter(date__gte=month_start, status='P').count()
    monthly_absent = teacher.attendances.filter(date__gte=month_start, status='A').count()
    monthly_late = teacher.attendances.filter(date__gte=month_start, status='L').count()
    monthly_excused = teacher.attendances.filter(date__gte=month_start, status='E').count()
    
    # All-time attendance totals
    total_present = teacher.attendances.filter(status='P').count()
    total_absent = teacher.attendances.filter(status='A').count()
    total_late = teacher.attendances.filter(status='L').count()
    total_excused = teacher.attendances.filter(status='E').count()
    
    # Count total students across all homeroom classes
    total_students = sum(c.students.count() for c in classes)
    
    # Get exams for subjects taught by this teacher
    teacher_subjects_ids = subjects.values_list('pk', flat=True)
    teacher_exams = Exam.objects.filter(subject_id__in=teacher_subjects_ids).select_related('subject', 'classroom', 'academic_year', 'exam_type').order_by('-date')[:20]
    
    # Timetable data for weekly view
    from school.models import TimeSlot
    time_slots = TimeSlot.objects.all().order_by('start_time')
    weekdays = ['ច័ន្ទ', 'អង្គារ', 'ពុធ', 'ព្រហស្បតិ៍', 'សុក្រ']
    weekdays_short = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    
    # Count total timetable slots
    total_timetable = timetables.count()
    total_exams = teacher_exams.count()
    
    return render(request, 'school/teacher_detail.html', {
        'teacher': teacher, 
        'subjects': subjects,
        'classes': classes, 
        'attendances': attendances,
        'all_attendances': all_attendances,
        'timetables': timetables,
        'attendance_rate': attendance_rate,
        'total_students': total_students,
        'teacher_exams': teacher_exams,
        'time_slots': time_slots,
        'weekdays': weekdays,
        'weekdays_short': weekdays_short,
        'total_timetable': total_timetable,
        'total_exams': total_exams,
        'monthly_present': monthly_present,
        'monthly_absent': monthly_absent,
        'monthly_late': monthly_late,
        'monthly_excused': monthly_excused,
        'total_present': total_present,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_excused': total_excused,
        'role': getattr(request.user.profile, 'role', 'teacher') if hasattr(request.user, 'profile') else 'teacher',
    })

@admin_required
def teacher_add(request):
    form = TeacherForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        try:
            teacher = form.save()
            # Check if account was created
            if form.cleaned_data.get('create_account'):
                try:
                    profile = UserProfile.objects.get(teacher=teacher)
                    username = profile.user.username
                    # Show success message with login credentials
                    messages.success(request, f'គ្រូបានបន្ថែមរួច។ គណនី: {username} | ពាក្យសម្ងាត់: teacher123')
                except UserProfile.DoesNotExist:
                    messages.success(request, 'គ្រូបានបន្ថែមរួច។')
            else:
                messages.success(request, 'គ្រូបានបន្ថែមរួច។')
            return redirect('school:teacher_list')
        except Exception as e:
            import traceback, logging
            logging.getLogger(__name__).error("teacher_add failed: %s\n%s", e, traceback.format_exc())
            messages.error(request, f'បញ្ហា: {e}')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'បន្ថែមគ្រូ', 'back_url': reverse('school:teacher_list')
    })

@admin_required
def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    form = TeacherForm(request.POST or None, request.FILES or None, instance=teacher)
    if request.method == 'POST' and form.is_valid():
        try:
            form.save()
            messages.success(request, 'គ្រូបានកែប្រែ។')
            return redirect('school:teacher_detail', pk=pk)
        except Exception as e:
            import traceback, logging
            logging.getLogger(__name__).error("teacher_edit failed: %s\n%s", e, traceback.format_exc())
            messages.error(request, f'បញ្ហា: {e}')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'កែប្រែគ្រូ',
        'subtitle': f'ID: {teacher.teacher_id}',
        'back_url': reverse('school:teacher_list'),
    })

@admin_required
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        teacher.delete()
        messages.success(request, 'គ្រូបានលុប។')
        return redirect('school:teacher_list')
    return render(request, 'school/confirm_delete.html', {
        'object': teacher, 'title': 'លុបគ្រូ', 'back_url': reverse('school:teacher_list')
    })


@admin_or_teacher
def teacher_document_upload(request, teacher_pk):
    from .forms import TeacherDocumentForm
    from .models import TeacherDocument
    
    teacher = get_object_or_404(Teacher, pk=teacher_pk)
    
    if request.method == 'POST':
        form = TeacherDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.teacher = teacher
            document.uploaded_by = request.user
            document.save()
            messages.success(request, 'ឯកសារបានបញ្ចូលរួច។')
            return redirect('school:teacher_detail', pk=teacher_pk)
    else:
        form = TeacherDocumentForm()
    
    # Get all existing documents
    documents = TeacherDocument.objects.filter(teacher=teacher)
    
    return render(request, 'school/teacher_document_upload.html', {
        'form': form,
        'teacher': teacher,
        'documents': documents,
        'title': f'បន្ថែមឯកសារសម្រាប់ {teacher.first_name} {teacher.last_name}',
        'back_url': reverse('school:teacher_detail', args=[teacher_pk])
    })


@admin_required
def teacher_document_delete(request, pk):
    from .models import TeacherDocument
    document = get_object_or_404(TeacherDocument, pk=pk)
    teacher_pk = document.teacher.pk
    
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'ឯកសារបានលុប។')
        return redirect('school:teacher_detail', pk=teacher_pk)
    
    return render(request, 'school/confirm_delete.html', {
        'object': document,
        'title': 'លុបឯកសារ',
        'back_url': reverse('school:teacher_detail', args=[teacher_pk])
    })


# ══════════════════════════════════════════════
#  CLASSROOMS & SUBJECTS (Admin only)
# ══════════════════════════════════════════════

# ══════════════════════════════════════════════
#  ACADEMIC YEAR
# ══════════════════════════════════════════════
@admin_required
def academic_year_list(request):
    years = AcademicYear.objects.all()
    return render(request, 'school/academic_year_list.html', {'years': years})

@admin_required
def academic_year_add(request):
    form = AcademicYearForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'ឆ្នាំសិក្សាបានបន្ថែម។')
        return redirect('school:academic_year_list')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'បន្ថែមឆ្នាំសិក្សា',
        'back_url': reverse('school:academic_year_list')
    })

@admin_required
def academic_year_edit(request, pk):
    year = get_object_or_404(AcademicYear, pk=pk)
    form = AcademicYearForm(request.POST or None, instance=year)
    if form.is_valid():
        form.save()
        messages.success(request, 'ឆ្នាំសិក្សាបានកែប្រែ។')
        return redirect('school:academic_year_list')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'កែប្រែឆ្នាំសិក្សា',
        'back_url': reverse('school:academic_year_list')
    })

@admin_required
def academic_year_delete(request, pk):
    year = get_object_or_404(AcademicYear, pk=pk)
    if request.method == 'POST':
        year.delete()
        messages.success(request, 'ឆ្នាំសិក្សាបានលុប។')
        return redirect('school:academic_year_list')
    return render(request, 'school/confirm_delete.html', {
        'object': year, 'title': 'លុបឆ្នាំសិក្សា',
        'back_url': reverse('school:academic_year_list')
    })

@admin_required
def academic_year_set_active(request, pk):
    AcademicYear.objects.all().update(is_active=False)
    AcademicYear.objects.filter(pk=pk).update(is_active=True)
    messages.success(request, 'ឆ្នាំសិក្សាសកម្មបានកំណត់។')
    return redirect('school:academic_year_list')

@admin_required
def academic_year_generate(request):
    """Generate multiple academic years at once."""
    if request.method == 'POST':
        base_year = int(request.POST.get('base_year', timezone.now().year))
        count = int(request.POST.get('count', 5))
        count = max(1, min(20, count))  # limit 1-20
        
        created_count = 0
        existing_count = 0
        
        for i in range(count):
            year_str = f"{base_year + i}-{base_year + i + 1}"
            obj, created = AcademicYear.objects.get_or_create(
                year=year_str,
                defaults={'is_active': False}
            )
            if created:
                created_count += 1
            else:
                existing_count += 1
        
        # Set first year as active if no active year exists
        if not AcademicYear.objects.filter(is_active=True).exists():
            first_year = AcademicYear.objects.order_by('year').first()
            if first_year:
                first_year.is_active = True
                first_year.save()
        
        if created_count > 0:
            messages.success(request, f'បានបង្កើតឆ្នាំសិក្សា {created_count} ឆ្នាំ។')
        if existing_count > 0:
            messages.info(request, f'{existing_count} ឆ្នាំមានរួចហើយ។')
        
        return redirect('school:academic_year_list')
    
    return redirect('school:academic_year_list')

# ══════════════════════════════════════════════
@admin_or_teacher
def classroom_list(request):
    classrooms = Classroom.objects.select_related('grade','homeroom_teacher','academic_year').annotate(student_count=Count('students'))
    role = getattr(request.user.profile, 'role', 'admin') if hasattr(request.user, 'profile') else 'admin'
    return render(request, 'school/classroom_list.html', {'classrooms': classrooms, 'role': role})

@admin_required
def classroom_add(request):
    form = ClassroomForm(request.POST or None)
    if form.is_valid():
        try:
            form.save()
            messages.success(request, 'ថ្នាក់បានបន្ថែម។')
            return redirect('school:classroom_list')
        except Exception as e:
            import traceback, logging
            logging.getLogger(__name__).error("classroom_add failed: %s\n%s", e, traceback.format_exc())
            messages.error(request, f'បញ្ហា: {e}')
    return render(request, 'school/form.html', {'form': form, 'title': 'បន្ថែមថ្នាក់', 'back_url': reverse('school:classroom_list')})

@admin_required
def classroom_edit(request, pk):
    classroom = get_object_or_404(Classroom, pk=pk)
    form = ClassroomForm(request.POST or None, instance=classroom)
    if form.is_valid():
        form.save(); messages.success(request, 'ថ្នាក់បានកែប្រែ។')
        return redirect('school:classroom_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែថ្នាក់', 'back_url': reverse('school:classroom_list')})

@admin_required
def classroom_delete(request, pk):
    classroom = get_object_or_404(Classroom, pk=pk)
    if request.method == 'POST':
        classroom.delete(); messages.success(request, 'ថ្នាក់បានលុប។')
        return redirect('school:classroom_list')
    return render(request, 'school/confirm_delete.html', {'object': classroom, 'title': 'លុបថ្នាក់', 'back_url': reverse('school:classroom_list')})

@admin_or_teacher
def subject_list(request):
    subjects = Subject.objects.select_related('teacher','grade')
    role = getattr(request.user.profile, 'role', 'admin') if hasattr(request.user, 'profile') else 'admin'
    return render(request, 'school/subject_list.html', {'subjects': subjects, 'role': role})

@admin_required
def subject_add(request):
    form = SubjectForm(request.POST or None)
    if form.is_valid():
        form.save(); messages.success(request, 'មុខវិជ្ជាបានបន្ថែម។')
        return redirect('school:subject_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'បន្ថែមមុខវិជ្ជា', 'back_url': reverse('school:subject_list')})

@admin_required
def subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    form = SubjectForm(request.POST or None, instance=subject)
    if form.is_valid():
        form.save(); messages.success(request, 'មុខវិជ្ជាបានកែប្រែ។')
        return redirect('school:subject_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែមុខវិជ្ជា', 'back_url': reverse('school:subject_list')})

@admin_required
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete(); messages.success(request, 'មុខវិជ្ជាបានលុប។')
        return redirect('school:subject_list')
    return render(request, 'school/confirm_delete.html', {'object': subject, 'title': 'លុបមុខវិជ្ជា', 'back_url': reverse('school:subject_list')})

# ══════════════════════════════════════════════
#  ATTENDANCE (Admin+Teacher: record | Parent+Student: view own)
# ══════════════════════════════════════════════
@admin_or_teacher
def attendance_list(request):
    today        = timezone.now().date()
    date_filter  = request.GET.get('date', str(today))
    classroom_id = request.GET.get('classroom', '')
    try:
        role = request.user.profile.role
    except Exception:
        role = 'teacher'
    records = Attendance.objects.filter(date=date_filter).select_related('student__classroom__grade')
    if role == 'teacher':
        try:
            teacher    = request.user.profile.teacher
            my_classes = Classroom.objects.filter(homeroom_teacher=teacher)
            records    = records.filter(student__classroom__in=my_classes)
        except Exception:
            records = Attendance.objects.none()
    elif classroom_id:
        records = records.filter(student__classroom_id=classroom_id)
    classrooms = Classroom.objects.select_related('grade','academic_year')
    return render(request, 'school/attendance_list.html', {
        'records': records, 'date_filter': date_filter,
        'classrooms': classrooms, 'selected_classroom': classroom_id,
        'present': records.filter(status='P').count(),
        'absent':  records.filter(status='A').count(),
        'late':    records.filter(status='L').count(),
        'role': role,
    })

@admin_or_teacher
def attendance_add(request):
    form = AttendanceForm(request.POST or None)
    if form.is_valid():
        form.save(); messages.success(request, 'វត្តមានបានកត់ត្រា។')
        return redirect('school:attendance_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កត់ត្រាវត្តមាន', 'back_url': reverse('school:attendance_list')})

@admin_or_teacher
def attendance_bulk(request):
    classroom = None
    students  = []
    date_val  = timezone.now().date()
    if request.method == 'POST':
        classroom_id = request.POST.get('classroom')
        date_val     = request.POST.get('date')
        classroom    = get_object_or_404(Classroom, pk=classroom_id)
        students     = classroom.students.filter(is_active=True).order_by('last_name')
        if 'save_attendance' in request.POST:
            for student in students:
                status = request.POST.get(f'status_{student.pk}', 'P')
                note   = request.POST.get(f'note_{student.pk}', '')
                Attendance.objects.update_or_create(
                    student=student, date=date_val,
                    defaults={'status': status, 'note': note}
                )
            messages.success(request, f'វត្តមានបានរក្សាទុក — {classroom} — {date_val}')
            return redirect('school:attendance_list')
    classrooms = Classroom.objects.select_related('grade','academic_year')
    return render(request, 'school/attendance_bulk.html', {
        'classrooms': classrooms, 'classroom': classroom,
        'students': students, 'date_val': date_val,
        'status_choices': Attendance.STATUS_CHOICES,
    })

# ── Parent view child attendance ───────────────
@role_required('parent')
def parent_child_attendance(request):
    try:
        student = request.user.profile.student
    except Exception:
        student = None
    attendances = student.attendances.order_by('-date') if student else []
    present = student.attendances.filter(status='P').count() if student else 0
    absent  = student.attendances.filter(status='A').count() if student else 0
    return render(request, 'school/parent/child_attendance.html', {
        'student': student, 'attendances': attendances,
        'present': present, 'absent': absent,
    })

# ── Student view own attendance ────────────────
@role_required('student')
def student_my_attendance(request):
    try:
        student = request.user.profile.student
    except Exception:
        student = None
    attendances = student.attendances.order_by('-date') if student else []
    return render(request, 'school/student/my_attendance.html', {
        'student': student, 'attendances': attendances,
    })


# ══════════════════════════════════════════════
#  TEACHER ATTENDANCE (Admin: full | Teacher: view own)
# ══════════════════════════════════════════════
@admin_or_teacher
def teacher_attendance_list(request):
    today       = timezone.now().date()
    date_filter = request.GET.get('date', str(today))
    try:
        role = request.user.profile.role
    except Exception:
        role = 'teacher'
    records = TeacherAttendance.objects.filter(date=date_filter).select_related('teacher')
    if role == 'teacher':
        try:
            teacher = request.user.profile.teacher
            records = records.filter(teacher=teacher)
        except Exception:
            records = TeacherAttendance.objects.none()
    return render(request, 'school/teacher_attendance_list.html', {
        'records':      records,
        'date_filter':  date_filter,
        'present':      records.filter(status='P').count(),
        'absent':       records.filter(status='A').count(),
        'late':         records.filter(status='L').count(),
        'role':         role,
    })


@admin_required
def teacher_attendance_add(request):
    form = TeacherAttendanceForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'វត្តមានគ្រូបានកត់ត្រា។')
        return redirect('school:teacher_attendance_list')
    return render(request, 'school/form.html', {
        'form': form, 'title': 'កត់ត្រាវត្តមានគ្រូ',
        'back_url': reverse('school:teacher_attendance_list')
    })


@admin_required
def teacher_attendance_bulk(request):
    """Record attendance for all active teachers at once."""
    date_val = timezone.now().date()
    teachers = []
    if request.method == 'POST':
        date_val = request.POST.get('date', str(date_val))
        if 'save_attendance' in request.POST:
            active_teachers = Teacher.objects.filter(is_active=True).order_by('last_name', 'first_name')
            for teacher in active_teachers:
                status = request.POST.get(f'status_{teacher.pk}', 'P')
                note   = request.POST.get(f'note_{teacher.pk}', '')
                TeacherAttendance.objects.update_or_create(
                    teacher=teacher, date=date_val,
                    defaults={'status': status, 'note': note}
                )
            messages.success(request, f'វត្តមានគ្រូបានរក្សាទុក — {date_val}')
            return redirect('school:teacher_attendance_list')
        else:
            teachers = Teacher.objects.filter(is_active=True).order_by('last_name', 'first_name')
    return render(request, 'school/teacher_attendance_bulk.html', {
        'teachers':       teachers,
        'date_val':       date_val,
        'status_choices': TeacherAttendance.STATUS_CHOICES,
    })

# ══════════════════════════════════════════════
#  EXAMS & SCORES (Admin+Teacher: manage | Parent+Student: view)
# ══════════════════════════════════════════════
@admin_or_teacher
def exam_list(request):
    exams = Exam.objects.select_related('subject','classroom','exam_type','academic_year').order_by('-date')
    role = getattr(request.user.profile, 'role', 'teacher') if hasattr(request.user, 'profile') else 'teacher'
    return render(request, 'school/exam_list.html', {'exams': exams, 'role': role})

@admin_or_teacher
def exam_add(request):
    form = ExamForm(request.POST or None)
    if form.is_valid():
        form.save(); messages.success(request, 'ការប្រឡងបានបន្ថែម។')
        return redirect('school:exam_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'បន្ថែមការប្រឡង', 'back_url': reverse('school:exam_list')})

@admin_or_teacher
def exam_edit(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    form = ExamForm(request.POST or None, instance=exam)
    if form.is_valid():
        form.save(); messages.success(request, 'ការប្រឡងបានកែប្រែ។')
        return redirect('school:exam_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែការប្រឡង', 'back_url': reverse('school:exam_list')})

@admin_required
def exam_delete(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        exam.delete(); messages.success(request, 'ការប្រឡងបានលុប។')
        return redirect('school:exam_list')
    return render(request, 'school/confirm_delete.html', {'object': exam, 'title': 'លុបការប្រឡង', 'back_url': reverse('school:exam_list')})

@admin_required
def exam_bulk_delete(request):
    if request.method == 'POST':
        exam_ids = request.POST.getlist('exam_ids')
        if exam_ids:
            deleted_count = Exam.objects.filter(pk__in=exam_ids).delete()[0]
            messages.success(request, f'បានលុបការប្រឡង {deleted_count} ដោយជោគជ័យ។')
        else:
            messages.warning(request, 'មិនមានការប្រឡងដែលបានជ្រើសរើស។')
    return redirect('school:exam_list')

@admin_or_teacher
def score_list(request):
    from django.db.models import Avg, Count
    try:
        role = request.user.profile.role
    except Exception:
        role = 'teacher'
    q    = request.GET.get('q','')
    scores = Score.objects.select_related('student','subject','exam_type','academic_year')
    if role == 'teacher':
        try:
            teacher    = request.user.profile.teacher
            my_classes = Classroom.objects.filter(homeroom_teacher=teacher)
            scores     = scores.filter(student__classroom__in=my_classes)
        except Exception:
            scores = Score.objects.none()
    if q:
        scores = scores.filter(Q(student__first_name__icontains=q)|Q(student__last_name__icontains=q)|Q(student__student_id__icontains=q))
    
    # Calculate statistics
    stats = {
        'total_scores': scores.count(),
        'avg_score': 0,
        'avg_percentage': 0,
        'pass_count': 0,
        'fail_count': 0,
    }
    
    if scores.exists():
        # Calculate averages
        total_percentage = sum(score.percentage() for score in scores)
        stats['avg_percentage'] = round(total_percentage / scores.count(), 1) if scores.count() > 0 else 0
        
        # Calculate pass/fail counts
        stats['pass_count'] = sum(1 for score in scores if score.is_passing(50))
        stats['fail_count'] = scores.count() - stats['pass_count']
    
    return render(request, 'school/score_list.html', {
        'scores': scores, 
        'q': q, 
        'role': role,
        'stats': stats,
    })

@admin_or_teacher
def score_add(request):
    form = ScoreForm(request.POST or None)
    if form.is_valid():
        form.save(); messages.success(request, 'ពិន្ទុបានកត់ត្រា។')
        return redirect('school:score_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'បន្ថែមពិន្ទុ', 'back_url': reverse('school:score_list')})

@admin_or_teacher
def score_edit(request, pk):
    score = get_object_or_404(Score, pk=pk)
    form  = ScoreForm(request.POST or None, instance=score)
    if form.is_valid():
        form.save(); messages.success(request, 'ពិន្ទុបានកែប្រែ។')
        return redirect('school:score_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែពិន្ទុ', 'back_url': reverse('school:score_list')})

@admin_or_teacher
def score_delete(request, pk):
    score = get_object_or_404(Score, pk=pk)
    if request.method == 'POST':
        score.delete(); messages.success(request, 'ពិន្ទុបានលុប។')
        return redirect('school:score_list')
    return render(request, 'school/confirm_delete.html', {'object': score, 'title': 'លុបពិន្ទុ', 'back_url': reverse('school:score_list')})

@admin_or_teacher
def score_bulk_entry(request):
    """Bulk score entry for a class by selecting an exam"""
    academic_years = AcademicYear.objects.all()
    classrooms = Classroom.objects.all()
    
    selected_year = request.GET.get('academic_year', '')
    selected_classroom = request.GET.get('classroom', '')
    selected_exam = request.GET.get('exam', '')
    
    # Filter exams based on selections
    exams = Exam.objects.none()
    if selected_year and selected_classroom:
        exams = Exam.objects.filter(
            academic_year_id=selected_year,
            classroom_id=selected_classroom
        ).select_related('subject', 'exam_type', 'classroom', 'academic_year')
    
    # Get exam and students if exam is selected
    exam = None
    students = []
    if selected_exam:
        exam = get_object_or_404(Exam, pk=selected_exam)
        students = exam.classroom.students.filter(is_active=True).order_by('student_id')
    
    # Handle POST - save scores
    if request.method == 'POST' and exam:
        student_ids = request.POST.getlist('student_ids')
        scores_list = request.POST.getlist('scores')
        remarks_list = request.POST.getlist('remarks')
        
        success_count = 0
        error_count = 0
        
        for i, student_id in enumerate(student_ids):
            score_value = scores_list[i] if i < len(scores_list) else ''
            remark = remarks_list[i] if i < len(remarks_list) else ''
            
            # Skip if no score entered
            if not score_value:
                continue
            
            try:
                student = Student.objects.get(pk=student_id)
                
                # Create or update score
                score_obj, created = Score.objects.update_or_create(
                    student=student,
                    subject=exam.subject,
                    exam_type=exam.exam_type,
                    academic_year=exam.academic_year,
                    defaults={
                        'exam': exam,
                        'score': float(score_value),
                        'max_score': exam.max_score,
                        'remarks': remark
                    }
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                import logging
                logging.getLogger(__name__).error(f"Failed to save score for student {student_id}: {e}")
        
        if success_count > 0:
            messages.success(request, f'បានរក្សាទុកពិន្ទុសម្រាប់សិស្ស {success_count} នាក់។')
        if error_count > 0:
            messages.warning(request, f'មានបញ្ហាក្នុងការរក្សាទុកពិន្ទុសម្រាប់សិស្ស {error_count} នាក់។')
        
        return redirect('school:score_list')
    
    # Safely get role
    try:
        role = request.user.profile.role
    except Exception:
        role = 'admin' if (request.user.is_superuser or request.user.is_staff) else 'student'
    
    return render(request, 'school/score_bulk_entry.html', {
        'academic_years': academic_years,
        'classrooms': classrooms,
        'exams': exams,
        'exam': exam,
        'students': students,
        'selected_year': selected_year,
        'selected_classroom': selected_classroom,
        'role': role,
    })


@admin_or_teacher
def score_multi_subject_entry(request):
    """
    Enter scores for multiple subjects (up to 4) for one student at once
    បញ្ចូលពិន្ទុច្រើនមុខវិជ្ជា (រហូតដល់ ៤ មុខ) សម្រាប់សិស្សម្នាក់តែម្តង
    """
    from .forms import BulkScoreEntryForm
    
    if request.method == 'POST':
        form = BulkScoreEntryForm(request.POST)
        if form.is_valid():
            student = form.cleaned_data['student']
            exam_type = form.cleaned_data['exam_type']
            academic_year = form.cleaned_data['academic_year']
            max_score = form.cleaned_data['max_score']
            
            success_count = 0
            errors = []
            
            # Process each subject (1-4)
            for i in range(1, 5):
                subject = form.cleaned_data.get(f'subject_{i}')
                score_value = form.cleaned_data.get(f'score_{i}')
                remarks = form.cleaned_data.get(f'remarks_{i}', '')
                
                # Skip if subject or score not provided
                if not subject or score_value is None:
                    continue
                
                try:
                    # Create or update score
                    score_obj, created = Score.objects.update_or_create(
                        student=student,
                        subject=subject,
                        exam_type=exam_type,
                        academic_year=academic_year,
                        defaults={
                            'score': score_value,
                            'max_score': max_score,
                            'remarks': remarks
                        }
                    )
                    success_count += 1
                except Exception as e:
                    errors.append(f"មុខវិជ្ជា {subject.name}: {str(e)}")
                    import logging
                    logging.getLogger(__name__).error(f"Failed to save score: {e}")
            
            # Display results
            if success_count > 0:
                messages.success(
                    request, 
                    f'✅ បានរក្សាទុកពិន្ទុ {success_count} មុខវិជ្ជាសម្រាប់សិស្ស {student.first_name} {student.last_name}'
                )
            
            if errors:
                for error in errors:
                    messages.error(request, f'❌ {error}')
            
            if success_count > 0:
                return redirect('school:score_list')
    else:
        form = BulkScoreEntryForm()
    
    # Safely get role
    try:
        role = request.user.profile.role
    except Exception:
        role = 'admin' if (request.user.is_superuser or request.user.is_staff) else 'student'
    
    return render(request, 'school/score_multi_subject_entry.html', {
        'form': form,
        'role': role,
        'title': 'បញ្ចូលពិន្ទុច្រើនមុខវិជ្ជា (Multi-Subject Score Entry)'
    })


@admin_or_teacher
def score_grid_entry(request):
    """
    Excel-style grade book: Enter scores for multiple students and subjects at once
    សៀវភៅពិន្ទុបែបតារាង៖ បញ្ចូលពិន្ទុច្រើននាក់និងច្រើនមុខវិជ្ជាតែម្តង
    """
    from django.http import JsonResponse
    
    # Get filter parameters
    classroom_id = request.GET.get('classroom', '')
    academic_year_id = request.GET.get('academic_year', '')
    exam_type_id = request.GET.get('exam_type', '')
    max_score = float(request.GET.get('max_score', 100))
    
    # Get all options for filters
    classrooms = Classroom.objects.select_related('grade', 'academic_year').filter(academic_year__is_active=True)
    academic_years = AcademicYear.objects.all()
    exam_types = ExamType.objects.all()
    
    students = []
    subjects = []
    
    # Load data if all filters are selected
    if classroom_id and academic_year_id and exam_type_id:
        classroom = get_object_or_404(Classroom, pk=classroom_id)
        academic_year = get_object_or_404(AcademicYear, pk=academic_year_id)
        exam_type = get_object_or_404(ExamType, pk=exam_type_id)
        
        # Get students in classroom
        students = classroom.students.filter(is_active=True).order_by('student_id')
        
        # Get subjects for this grade
        subjects = Subject.objects.filter(grade=classroom.grade).order_by('name')
        
        # Load existing scores
        existing_scores = {}
        if students and subjects:
            scores = Score.objects.filter(
                student__in=students,
                subject__in=subjects,
                exam_type=exam_type,
                academic_year=academic_year
            ).select_related('student', 'subject')
            
            for score in scores:
                key = f"{score.student.id}_{score.subject.id}"
                existing_scores[key] = score.score
    
    # Handle POST - Save scores
    if request.method == 'POST':
        saved_count = 0
        errors = []
        
        for key, value in request.POST.items():
            if key.startswith('score_'):
                # Parse key: score_studentid_subjectid
                try:
                    parts = key.split('_')
                    if len(parts) != 3:
                        continue
                    
                    student_id = int(parts[1])
                    subject_id = int(parts[2])
                    score_value = float(value) if value else None
                    
                    if score_value is None or score_value < 0:
                        continue
                    
                    student = Student.objects.get(pk=student_id)
                    subject = Subject.objects.get(pk=subject_id)
                    exam_type = ExamType.objects.get(pk=exam_type_id)
                    academic_year = AcademicYear.objects.get(pk=academic_year_id)
                    
                    # Create or update score
                    Score.objects.update_or_create(
                        student=student,
                        subject=subject,
                        exam_type=exam_type,
                        academic_year=academic_year,
                        defaults={
                            'score': score_value,
                            'max_score': max_score
                        }
                    )
                    saved_count += 1
                    
                except Exception as e:
                    errors.append(str(e))
                    import logging
                    logging.getLogger(__name__).error(f"Failed to save score {key}: {e}")
        
        # Return JSON response
        return JsonResponse({
            'success': True,
            'saved_count': saved_count,
            'errors': errors
        })
    
    # Safely get role
    try:
        role = request.user.profile.role
    except Exception:
        role = 'admin' if (request.user.is_superuser or request.user.is_staff) else 'student'
    
    return render(request, 'school/score_grid_entry.html', {
        'classrooms': classrooms,
        'academic_years': academic_years,
        'exam_types': exam_types,
        'students': students,
        'subjects': subjects,
        'classroom_id': classroom_id,
        'academic_year_id': academic_year_id,
        'exam_type_id': exam_type_id,
        'max_score': max_score,
        'existing_scores': existing_scores if 'existing_scores' in locals() else {},
        'role': role,
    })


# ── Parent/Student view results ────────────────
@role_required('parent')
def parent_child_results(request):
    try:
        student = request.user.profile.student
    except Exception:
        student = None
    scores = student.scores.select_related('subject','exam_type','academic_year').order_by('subject__name') if student else []
    
    # Calculate statistics
    stats = {
        'total_scores': len(scores),
        'avg_percentage': 0,
        'pass_count': 0,
        'fail_count': 0,
        'highest_score': 0,
        'lowest_score': 0,
    }
    
    if scores:
        percentages = [score.percentage() for score in scores]
        stats['avg_percentage'] = round(sum(percentages) / len(percentages), 1) if percentages else 0
        stats['pass_count'] = sum(1 for score in scores if score.is_passing(50))
        stats['fail_count'] = len(scores) - stats['pass_count']
        stats['highest_score'] = max(percentages) if percentages else 0
        stats['lowest_score'] = min(percentages) if percentages else 0
    
    return render(request, 'school/parent/child_results.html', {
        'student': student, 
        'scores': scores,
        'stats': stats,
    })

@role_required('student')
def student_my_results(request):
    try:
        student = request.user.profile.student
    except Exception:
        student = None
    scores = student.scores.select_related('subject','exam_type','academic_year').order_by('subject__name') if student else []
    
    # Calculate statistics
    stats = {
        'total_scores': len(scores),
        'avg_percentage': 0,
        'pass_count': 0,
        'fail_count': 0,
        'highest_score': 0,
        'lowest_score': 0,
    }
    
    if scores:
        percentages = [score.percentage() for score in scores]
        stats['avg_percentage'] = round(sum(percentages) / len(percentages), 1) if percentages else 0
        stats['pass_count'] = sum(1 for score in scores if score.is_passing(50))
        stats['fail_count'] = len(scores) - stats['pass_count']
        stats['highest_score'] = max(percentages) if percentages else 0
        stats['lowest_score'] = min(percentages) if percentages else 0
    
    return render(request, 'school/student/my_results.html', {
        'student': student, 
        'scores': scores,
        'stats': stats,
    })

# ══════════════════════════════════════════════
#  TIMETABLE (Admin: manage | Teacher+Student: view)
# ══════════════════════════════════════════════
@login_required
def timetable_list(request):
    """
    Display timetable in Cambodia school weekly grid format
    បង្ហាញកាលវិភាគជាទ្រង់ទ្រាយតារាងសប្តាហ៍ប្រភេទសាលាកម្ពុជា
    """
    try:
        role = request.user.profile.role
    except Exception:
        role = 'student'
    
    classroom_id = request.GET.get('classroom', '')
    year_id = request.GET.get('year', '')
    
    classrooms = Classroom.objects.select_related('grade', 'academic_year')
    years = AcademicYear.objects.all()
    
    # Get selected classroom object for display
    selected_classroom_obj = None
    if classroom_id:
        try:
            selected_classroom_obj = Classroom.objects.get(pk=classroom_id)
        except Classroom.DoesNotExist:
            pass
    
    timetables = Timetable.objects.select_related(
        'subject', 'teacher', 'time_slot', 'classroom__grade', 'academic_year'
    ).order_by('time_slot__day', 'time_slot__period')
    
    # Student auto-filter to their class
    if role == 'student':
        try:
            student = request.user.profile.student
            if student and student.classroom:
                timetables = timetables.filter(classroom=student.classroom)
                selected_classroom_obj = student.classroom
        except Exception:
            pass
    else:
        if classroom_id:
            timetables = timetables.filter(classroom_id=classroom_id)
        if year_id:
            timetables = timetables.filter(academic_year_id=year_id)
    
    # Organize by day for grid view
    days = {}
    for tt in timetables:
        days.setdefault(tt.time_slot.day, []).append(tt)
    
    day_names = {
        1: 'ច័ន្ទ', 
        2: 'អង្គារ', 
        3: 'ពុធ', 
        4: 'ព្រហស្បតិ៍', 
        5: 'សុក្រ', 
        6: 'សៅរ៍'
    }
    
    # Cambodia timetable structure
    periods = [1, 2, 'break1', 3, 4, 5, 6, 7, 'break2', 8, 9]
    period_times = {
        1: '07:00-07:50',
        2: '07:50-08:40',
        3: '08:55-09:45',
        4: '09:45-10:35',
        5: '10:35-11:25',
        6: '13:30-14:20',
        7: '14:20-15:10',
        8: '15:25-16:15',
        9: '16:15-17:05',
    }
    
    return render(request, 'school/timetable_list.html', {
        'timetables': timetables,
        'classrooms': classrooms,
        'years': years,
        'days': days,
        'day_names': day_names,
        'periods': periods,
        'period_times': period_times,
        'selected_classroom': classroom_id,
        'selected_classroom_obj': selected_classroom_obj,
        'selected_year': year_id,
        'role': role,
    })

@admin_required
def timetable_add(request):
    form = TimetableForm(request.POST or None)
    if form.is_valid():
        form.save(); messages.success(request, 'តារាងម៉ោងបានបន្ថែម។')
        return redirect('school:timetable_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'បន្ថែមតារាងម៉ោង', 'back_url': reverse('school:timetable_list')})

@admin_required
def timetable_copy(request):
    """
    Copy timetable from one classroom/year to another
    ចម្លងកាលវិភាគពីថ្នាក់មួយទៅថ្នាក់ផ្សេង
    """
    if request.method == 'POST':
        source_classroom_id = request.POST.get('source_classroom')
        target_classroom_id = request.POST.get('target_classroom')
        target_academic_year_id = request.POST.get('target_academic_year')
        
        if source_classroom_id and target_classroom_id:
            source_classroom = get_object_or_404(Classroom, pk=source_classroom_id)
            target_classroom = get_object_or_404(Classroom, pk=target_classroom_id)
            
            # Use target academic year if specified, otherwise use target classroom's year
            if target_academic_year_id:
                target_year = get_object_or_404(AcademicYear, pk=target_academic_year_id)
            else:
                target_year = target_classroom.academic_year
            
            # Get source timetables
            source_timetables = Timetable.objects.filter(classroom=source_classroom)
            
            if not source_timetables.exists():
                messages.warning(request, f'⚠️ ថ្នាក់ {source_classroom} មិនមានកាលវិភាគទេ។')
                return redirect('school:timetable_copy')
            
            # Delete existing timetables for target if requested
            replace_existing = request.POST.get('replace_existing') == 'yes'
            if replace_existing:
                deleted_count = Timetable.objects.filter(
                    classroom=target_classroom,
                    academic_year=target_year
                ).delete()[0]
                if deleted_count > 0:
                    messages.info(request, f'🗑️ បានលុបកាលវិភាគចាស់ {deleted_count} ធាតុ។')
            
            # Copy timetables
            copied_count = 0
            skipped_count = 0
            
            for source_tt in source_timetables:
                # Check if already exists (to avoid duplicates)
                exists = Timetable.objects.filter(
                    classroom=target_classroom,
                    time_slot=source_tt.time_slot,
                    academic_year=target_year
                ).exists()
                
                if not exists:
                    Timetable.objects.create(
                        classroom=target_classroom,
                        subject=source_tt.subject,
                        teacher=source_tt.teacher,
                        time_slot=source_tt.time_slot,
                        academic_year=target_year,
                        room=source_tt.room
                    )
                    copied_count += 1
                else:
                    skipped_count += 1
            
            if copied_count > 0:
                messages.success(request, f'✅ បានចម្លងកាលវិភាគ {copied_count} ធាតុទៅ {target_classroom} ({target_year})។')
            if skipped_count > 0:
                messages.info(request, f'ℹ️ រំលងធាតុដែលមានស្រាប់ {skipped_count} ធាតុ។')
            
            return redirect('school:timetable_list')
    
    # GET request - show form
    classrooms = Classroom.objects.select_related('grade', 'academic_year').order_by('grade__grade_number')
    academic_years = AcademicYear.objects.all()
    
    # Add timetable count to classrooms
    classrooms_with_count = []
    for classroom in classrooms:
        count = classroom.timetables.count()
        classrooms_with_count.append({
            'classroom': classroom,
            'timetable_count': count
        })
    
    return render(request, 'school/timetable_copy.html', {
        'classrooms_with_count': classrooms_with_count,
        'academic_years': academic_years,
    })

@admin_required
def timetable_edit(request, pk):
    tt   = get_object_or_404(Timetable, pk=pk)
    form = TimetableForm(request.POST or None, instance=tt)
    if form.is_valid():
        form.save(); messages.success(request, 'តារាងម៉ោងបានកែប្រែ។')
        return redirect('school:timetable_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែតារាងម៉ោង', 'back_url': reverse('school:timetable_list')})

@admin_required
def timetable_delete(request, pk):
    tt = get_object_or_404(Timetable, pk=pk)
    if request.method == 'POST':
        tt.delete(); messages.success(request, 'តារាងម៉ោងបានលុប។')
        return redirect('school:timetable_list')
    return render(request, 'school/confirm_delete.html', {'object': tt, 'title': 'លុបតារាងម៉ោង', 'back_url': reverse('school:timetable_list')})

# ══════════════════════════════════════════════
#  NOTIFICATIONS (Admin+Teacher: manage | All: receive)
# ══════════════════════════════════════════════
@login_required
def notification_list(request):
    try:
        role = request.user.profile.role
    except Exception:
        role = 'student'
    notifs = Notification.objects.select_related('created_by').filter(is_active=True)
    # Filter by audience
    if role == 'teacher':
        notifs = notifs.filter(audience__in=['all','teachers'])
    elif role == 'parent':
        notifs = notifs.filter(audience__in=['all','parents'])
    elif role == 'student':
        notifs = notifs.filter(audience__in=['all','students'])
    return render(request, 'school/notification_list.html', {
        'notifications': notifs.order_by('-created_at'), 'role': role
    })

@admin_or_teacher
def notification_add(request):
    form = NotificationForm(request.POST or None)
    if form.is_valid():
        notif = form.save(commit=False)
        notif.created_by = request.user
        notif.save()
        messages.success(request, 'សេចក្ដីជូនដំណឹងបានផ្ញើ។')
        return redirect('school:notification_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'ផ្ញើសេចក្ដីជូនដំណឹង', 'back_url': reverse('school:notification_list')})

@admin_or_teacher
def notification_edit(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    form  = NotificationForm(request.POST or None, instance=notif)
    if form.is_valid():
        form.save(); messages.success(request, 'សេចក្ដីជូនដំណឹងបានកែប្រែ។')
        return redirect('school:notification_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែ', 'back_url': reverse('school:notification_list')})

@admin_required
def notification_delete(request, pk):
    notif = get_object_or_404(Notification, pk=pk)
    if request.method == 'POST':
        notif.delete(); messages.success(request, 'បានលុប។')
        return redirect('school:notification_list')
    return render(request, 'school/confirm_delete.html', {'object': notif, 'title': 'លុបសេចក្ដីជូនដំណឹង', 'back_url': reverse('school:notification_list')})

# ══════════════════════════════════════════════
#  EVENTS (Admin: manage | All: view)
# ══════════════════════════════════════════════
@login_required
def event_list(request):
    events = SchoolEvent.objects.order_by('start_date')
    role = getattr(request.user.profile, 'role', 'student') if hasattr(request.user, 'profile') else 'student'
    return render(request, 'school/event_list.html', {'events': events, 'role': role})

@admin_required
def event_add(request):
    form = SchoolEventForm(request.POST or None)
    if form.is_valid():
        ev = form.save(commit=False); ev.created_by = request.user; ev.save()
        messages.success(request, 'ព្រឹត្តិការណ៍បានបន្ថែម។')
        return redirect('school:event_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'បន្ថែមព្រឹត្តិការណ៍', 'back_url': reverse('school:event_list')})

@admin_required
def event_edit(request, pk):
    ev   = get_object_or_404(SchoolEvent, pk=pk)
    form = SchoolEventForm(request.POST or None, instance=ev)
    if form.is_valid():
        form.save(); messages.success(request, 'ព្រឹត្តិការណ៍បានកែប្រែ។')
        return redirect('school:event_list')
    return render(request, 'school/form.html', {'form': form, 'title': 'កែប្រែព្រឹត្តិការណ៍', 'back_url': reverse('school:event_list')})

@admin_required
def event_delete(request, pk):
    ev = get_object_or_404(SchoolEvent, pk=pk)
    if request.method == 'POST':
        ev.delete(); messages.success(request, 'ព្រឹត្តិការណ៍បានលុប។')
        return redirect('school:event_list')
    return render(request, 'school/confirm_delete.html', {'object': ev, 'title': 'លុបព្រឹត្តិការណ៍', 'back_url': reverse('school:event_list')})

# ══════════════════════════════════════════════
#  REPORT CARDS (Admin+Teacher: manage | Parent+Student: view)
# ══════════════════════════════════════════════
@admin_or_teacher
def report_card_list(request):
    try:
        role = request.user.profile.role
    except Exception:
        role = 'teacher'
    cards = ReportCard.objects.select_related('student__classroom__grade','academic_year').order_by('-generated_at')
    if role == 'teacher':
        try:
            teacher    = request.user.profile.teacher
            my_classes = Classroom.objects.filter(homeroom_teacher=teacher)
            cards      = cards.filter(student__classroom__in=my_classes)
        except Exception:
            cards = ReportCard.objects.none()
    return render(request, 'school/report_card_list.html', {'cards': cards, 'role': role})

@admin_or_teacher
def report_card_add(request):
    # Get filter parameters
    classroom_id = request.GET.get('classroom', '')
    academic_year_id = request.GET.get('academic_year', '')
    term = request.GET.get('term', '')
    
    classrooms = Classroom.objects.select_related('grade', 'academic_year')
    academic_years = AcademicYear.objects.all()
    students_data = []
    
    # Load students if classroom is selected
    if classroom_id:
        classroom = get_object_or_404(Classroom, pk=classroom_id)
        students = Student.objects.filter(
            classroom=classroom,
            is_active=True
        ).order_by('last_name', 'first_name')
        
        # Get academic year
        if academic_year_id:
            academic_year = get_object_or_404(AcademicYear, pk=academic_year_id)
        else:
            academic_year = classroom.academic_year
        
        # Prepare student data with scores
        for student in students:
            scores = student.scores.filter(academic_year=academic_year) if academic_year else student.scores.all()
            avg_score = scores.aggregate(avg=Avg('score'))['avg'] or 0
            
            # Check if report card already exists
            existing_card = ReportCard.objects.filter(
                student=student,
                academic_year=academic_year,
                term=term
            ).first() if academic_year and term else None
            
            students_data.append({
                'student': student,
                'avg_score': round(avg_score, 2),
                'total_subjects': scores.count(),
                'existing_card': existing_card,
            })
    
    # Handle POST - create multiple report cards
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        next_academic_year_id = request.POST.get('academic_year')
        next_term = request.POST.get('term')
        status = request.POST.get('status', 'Draft')
        teacher_remarks = request.POST.get('teacher_remarks', '')
        principal_remarks = request.POST.get('principal_remarks', '')
        conduct = request.POST.get('conduct', '')
        attendance_days = int(request.POST.get('attendance_days', 0))
        absent_days = int(request.POST.get('absent_days', 0))
        
        if student_ids and next_academic_year_id and next_term:
            academic_year = get_object_or_404(AcademicYear, pk=next_academic_year_id)
            created_count = 0
            
            for student_id in student_ids:
                try:
                    student = Student.objects.get(pk=student_id)
                    
                    # Check if already exists
                    existing = ReportCard.objects.filter(
                        student=student,
                        academic_year=academic_year,
                        term=next_term
                    ).first()
                    
                    if not existing:
                        ReportCard.objects.create(
                            student=student,
                            academic_year=academic_year,
                            term=next_term,
                            status=status,
                            teacher_remarks=teacher_remarks,
                            principal_remarks=principal_remarks,
                            conduct=conduct,
                            attendance_days=attendance_days,
                            absent_days=absent_days,
                            generated_by=request.user
                        )
                        created_count += 1
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error(f"Failed to create report card for student {student_id}: {e}")
            
            if created_count > 0:
                messages.success(request, f'បានបង្កើតសៀវភៅប័ណ្ណ {created_count} ចំណុះ។')
            return redirect('school:report_card_list')
    
    return render(request, 'school/report_card_add.html', {
        'classrooms': classrooms,
        'academic_years': academic_years,
        'students_data': students_data,
        'classroom_id': classroom_id,
        'academic_year_id': academic_year_id,
        'term': term,
    })

@login_required
def report_card_view(request, pk):
    card    = get_object_or_404(ReportCard, pk=pk)
    # Parent/Student can only view their own child/self
    try:
        role = request.user.profile.role
    except Exception:
        role = 'student'
    if role == 'parent' or role == 'student':
        try:
            my_student = request.user.profile.student
            if card.student != my_student:
                messages.error(request, 'អ្នកមិនមានសិទ្ធិមើលប័ណ្ណនេះ។')
                return redirect('school:dashboard')
        except Exception:
            return redirect('school:dashboard')
    student      = card.student
    scores       = student.scores.filter(academic_year=card.academic_year).select_related('subject','exam_type').order_by('subject__name')
    present_days = card.attendance_days - card.absent_days
    avg_score    = scores.aggregate(avg=Avg('score'))['avg'] or 0
    return render(request, 'school/report_card_print.html', {
        'card': card, 'student': student, 'scores': scores,
        'total_att': card.attendance_days, 'absent_days': card.absent_days,
        'present_days': present_days, 'avg_score': round(avg_score, 2),
        'today': timezone.now().date(),
    })

@admin_required
def report_card_delete(request, pk):
    card = get_object_or_404(ReportCard, pk=pk)
    if request.method == 'POST':
        card.delete(); messages.success(request, 'សៀវភៅប័ណ្ណបានលុប។')
        return redirect('school:report_card_list')
    return render(request, 'school/confirm_delete.html', {'object': card, 'title': 'លុបសៀវភៅប័ណ្ណ', 'back_url': reverse('school:report_card_list')})

# ══════════════════════════════════════════════
#  PRINT REPORTS (Admin: all | Teacher: own class)
# ══════════════════════════════════════════════
@admin_or_teacher
def report_students(request):
    try:
        role = request.user.profile.role
    except Exception:
        role = 'teacher'
    students = Student.objects.filter(is_active=True).select_related('classroom__grade').order_by('student_id')
    if role == 'teacher':
        try:
            teacher    = request.user.profile.teacher
            my_classes = Classroom.objects.filter(homeroom_teacher=teacher)
            students   = students.filter(classroom__in=my_classes)
        except Exception:
            students = Student.objects.none()
    return render(request, 'school/reports/report_students.html', {'students': students, 'today': timezone.now().date()})

@admin_required
def report_teachers(request):
    teachers = Teacher.objects.filter(is_active=True).order_by('teacher_id')
    return render(request, 'school/reports/report_teachers.html', {'teachers': teachers, 'today': timezone.now().date()})

@admin_or_teacher
def report_attendance(request):
    today       = timezone.now().date()
    date_filter = request.GET.get('date', str(today))
    records     = Attendance.objects.filter(date=date_filter).select_related('student').order_by('student__student_id')
    return render(request, 'school/reports/report_attendance.html', {
        'records': records, 'date_filter': date_filter,
        'present': records.filter(status='P').count(),
        'absent':  records.filter(status='A').count(),
        'late':    records.filter(status='L').count(),
        'excused': records.filter(status='E').count(),
        'today': today,
    })

@admin_or_teacher
def report_scores(request):
    academic_year_id = request.GET.get('year','')
    years  = AcademicYear.objects.all()
    scores = Score.objects.select_related('student','subject','exam_type','academic_year')
    selected_year = None
    if academic_year_id:
        scores        = scores.filter(academic_year_id=academic_year_id)
        selected_year = AcademicYear.objects.filter(pk=academic_year_id).first()
    return render(request, 'school/reports/report_scores.html', {
        'scores': scores, 'years': years,
        'selected_year': selected_year, 'today': timezone.now().date(),
    })

@admin_or_teacher
def report_student_detail(request, pk):
    student       = get_object_or_404(Student, pk=pk)
    attendances   = student.attendances.order_by('-date')
    scores        = student.scores.select_related('subject','exam_type','academic_year').order_by('subject__name')
    present_count = attendances.filter(status='P').count()
    absent_count  = attendances.filter(status='A').count()
    return render(request, 'school/reports/report_student_detail.html', {
        'student': student, 'attendances': attendances, 'scores': scores,
        'present_count': present_count, 'absent_count': absent_count,
        'today': timezone.now().date(),
    })
